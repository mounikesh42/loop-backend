#!/usr/bin/env python3
"""Easy wrapper: Run the GCP pipeline and export all results to Excel sheets."""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import run_pipeline  # noqa: E402

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


def flatten_dict(d, parent_key="", sep="_"):
    """Flatten nested dictionaries."""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, (list, tuple)):
            # Convert lists to JSON string
            items.append((new_key, json.dumps(v)))
        else:
            items.append((new_key, v))
    return dict(items)


def export_json_to_excel(json_file, excel_file, sheet_name):
    """Convert a JSON envelope file to an Excel sheet."""
    with open(json_file, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    # Extract the envelope data
    envelope = data.get("envelope", data)
    
    if not isinstance(envelope, list):
        # Single object or nested structure - wrap in list
        envelope = [envelope]
    
    if not envelope:
        print(f"  ⊘ {json_file.name} (empty data, skipping)")
        return 0
    
    # Flatten all records
    flat_records = [flatten_dict(record) for record in envelope]
    row_count = len(flat_records)
    col_count = len(flat_records[0]) if flat_records else 0
    
    # Load or create workbook
    if excel_file.exists():
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
        # Remove default sheet if it exists
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    
    # Create or get sheet (sanitize name to max 31 chars for Excel)
    safe_sheet_name = sheet_name[:31]
    if safe_sheet_name in wb.sheetnames:
        ws = wb[safe_sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(safe_sheet_name)
    
    # Write header
    if flat_records:
        fieldnames = sorted(flat_records[0].keys())
        ws.append(fieldnames)
        
        # Write data
        for record in flat_records:
            row = [record.get(fn, "") for fn in fieldnames]
            ws.append(row)
    
    wb.save(excel_file)
    print(f"  ✓ {json_file.name} -> sheet '{safe_sheet_name}' ({row_count} rows, {col_count} cols)")
    return row_count


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Run GCP pipeline and export all outputs to Excel sheets",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python run_and_export.py paths.json
  python run_and_export.py paths.json --excel-file results.xlsx
  python run_and_export.py paths.json --excel-file gcp_results.xlsx --skip-pipeline
        """
    )
    parser.add_argument("config", help="Path to paths.json config file")
    parser.add_argument(
        "--excel-file",
        default="gcp_results.xlsx",
        help="Excel file to save results (each run creates a new sheet, default: gcp_results.xlsx)"
    )
    parser.add_argument(
        "--skip-pipeline",
        action="store_true",
        help="Skip pipeline run, only export existing outputs to Excel"
    )
    
    args = parser.parse_args(argv)
    
    if not HAS_OPENPYXL:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        return 1
    
    config_path = Path(args.config).resolve()
    if not config_path.exists():
        print(f"ERROR: config not found: {config_path}", file=sys.stderr)
        return 1
    
    excel_file = Path(args.excel_file).resolve()
    excel_file.parent.mkdir(parents=True, exist_ok=True)
    
    # Run pipeline
    if not args.skip_pipeline:
        print(f"\n📊 Running GCP pipeline...")
        result = run_pipeline.main([str(config_path)])
        if result != 0:
            print(f"ERROR: Pipeline failed with exit code {result}", file=sys.stderr)
            return result
        print(f"✓ Pipeline completed successfully\n")
    
    # Load config to find output files
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)
    
    root = config_path.parent
    outputs = config.get("outputs", {})
    
    # Determine sheet name (timestamp-based or run count)
    if excel_file.exists():
        try:
            wb = load_workbook(excel_file)
            existing_sheets = wb.sheetnames
            sheet_num = len(existing_sheets) + 1
            sheet_name = f"Run_{sheet_num}"
            wb.close()
        except Exception as e:
            print(f"ERROR: Could not read existing Excel file: {e}", file=sys.stderr)
            return 1
    else:
        sheet_name = "Run_1"
    
    # Export each output to Excel sheet
    print(f"📊 Exporting to Excel sheets ({excel_file}):")
    try:
        for key, rel_path in outputs.items():
            json_file = root / rel_path
            if not json_file.exists():
                print(f"  ⊘ {json_file.name} not found, skipping")
                continue
            
            export_json_to_excel(json_file, excel_file, f"{sheet_name}_{key[:20]}")
    except Exception as e:
        print(f"  ✗ ERROR during export: {e}", file=sys.stderr)
        return 1
    
    print(f"\n✅ Done! Results appended to: {excel_file.resolve()}")
    print(f"   Sheet name: '{sheet_name}' (with substage names)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
