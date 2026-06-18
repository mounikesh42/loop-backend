#!/usr/bin/env python3
"""CSV/Excel export utilities for pipeline results.

Provides functions to export pipeline stage data (especially final scores) as
CSV or Excel with custom output filenames.
"""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any


def flatten_contributions(contributions: list[dict]) -> list[dict]:
    """Flatten contribution data for CSV export."""
    return [
        {
            "block_id": c.get("block_id"),
            "block_name": c.get("block_name"),
            "block_score": c.get("block_score"),
            "contribution": c.get("contribution"),
            "weight_in_apex": c.get("weight_in_apex"),
        }
        for c in contributions
    ]


def flatten_flags(flags: list[dict]) -> list[dict]:
    """Flatten flag data for CSV export."""
    return [
        {
            "flag_id": f.get("flag_id"),
            "flag_name": f.get("flag_name"),
            "severity": f.get("severity"),
            "origin_stage": f.get("_origin_stage", ""),
            "indicator_id": f.get("_indicator_id", ""),
        }
        for f in flags
    ]


def flatten_per_deliverable_views(views: dict[str, Any]) -> list[dict]:
    """Flatten per-deliverable view scores for CSV export."""
    return [
        {
            "deliverable": k.replace("VIEW_", ""),
            "score": v,
        }
        for k, v in sorted(views.items())
    ]


def export_processing_score_csv(
    envelope: dict,
    output_path: Path,
) -> None:
    """Export final processing score (stage 3d) to CSV.

    Creates CSV with main score summary + separate rows for contributions.

    Args:
        envelope: The stage 3d envelope dict from run_pipeline
        output_path: Path to write CSV file to
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    data = envelope.get("data", {})
    config = envelope.get("config_used", {})

    # Main score summary row
    with output_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=[
                "survey_id",
                "subsystem",
                "spec_version",
                "generated_at",
                "processing_score",
                "verification_status",
                "is_null",
                "global_gate_triggered",
                "total_flags",
                "unique_flags",
                "critical_flags",
                "high_flags",
                "medium_flags",
                "informational_flags",
            ],
        )
        writer.writeheader()

        verification = data.get("verification_status", {})
        verification_value = verification.get("value", "UNKNOWN") if isinstance(verification, dict) else str(verification)
        flags_by_severity = data.get("flags_by_severity", {})

        writer.writerow(
            {
                "survey_id": config.get("survey_id", ""),
                "subsystem": config.get("subsystem", ""),
                "spec_version": envelope.get("spec_version", ""),
                "generated_at": envelope.get("generated_at", ""),
                "processing_score": data.get("processing_score", ""),
                "verification_status": verification_value,
                "is_null": data.get("null_handling", {}).get("is_null", False),
                "global_gate_triggered": data.get("global_gate", {}).get("triggered", False),
                "total_flags": data.get("stage3d_meta", {}).get("total_flags_aggregated", 0),
                "unique_flags": data.get("stage3d_meta", {}).get("unique_flag_count", 0),
                "critical_flags": flags_by_severity.get("CRITICAL", 0),
                "high_flags": flags_by_severity.get("HIGH", 0),
                "medium_flags": flags_by_severity.get("MEDIUM", 0),
                "informational_flags": flags_by_severity.get("INFORMATIONAL", 0),
            }
        )


def export_processing_score_xlsx(
    envelope: dict,
    output_path: Path,
) -> None:
    """Export final processing score to Excel with multiple sheets.

    Requires openpyxl. Creates sheets:
      - Summary: main score + verification status
      - Score_Contributions: block scores and weighted contributions
      - Per_Deliverable_Views: scores for each deliverable
      - Flags: all flags with severity and origin

    Args:
        envelope: The stage 3d envelope dict from run_pipeline
        output_path: Path to write Excel file to
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError(
            "openpyxl required for Excel export. Install with: pip install openpyxl"
        )

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    data = envelope.get("data", {})
    config = envelope.get("config_used", {})

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # ===== Sheet 1: Summary =====
    ws_summary["A1"] = "Processing Score Summary"
    ws_summary["A1"].font = Font(bold=True, size=12)

    verification = data.get("verification_status", {})
    verification_value = verification.get("value", "UNKNOWN") if isinstance(verification, dict) else str(verification)

    summary_rows = [
        ("Survey ID", config.get("survey_id", "")),
        ("Subsystem", config.get("subsystem", "")),
        ("Spec Version", envelope.get("spec_version", "")),
        ("Generated At", envelope.get("generated_at", "")),
        ("Processing Score", data.get("processing_score", "")),
        ("Verification Status", verification_value),
        ("Is Null", data.get("null_handling", {}).get("is_null", False)),
        ("Global Gate Triggered", data.get("global_gate", {}).get("triggered", False)),
    ]
    for idx, (key, value) in enumerate(summary_rows, start=2):
        ws_summary[f"A{idx}"] = key
        ws_summary[f"B{idx}"] = value
        ws_summary[f"A{idx}"].font = Font(bold=True)

    # ===== Sheet 2: Score Contributions =====
    ws_contrib = wb.create_sheet("Score_Contributions")
    contributions = flatten_contributions(data.get("contributions", []))

    contrib_headers = ["block_id", "block_name", "block_score", "contribution", "weight_in_apex"]
    for col_idx, header in enumerate(contrib_headers, start=1):
        ws_contrib.cell(row=1, column=col_idx, value=header)
        ws_contrib.cell(row=1, column=col_idx).font = Font(bold=True)

    for row_idx, contrib in enumerate(contributions, start=2):
        for col_idx, header in enumerate(contrib_headers, start=1):
            ws_contrib.cell(row=row_idx, column=col_idx, value=contrib.get(header))

    # ===== Sheet 3: Per-Deliverable Views =====
    ws_deliv = wb.create_sheet("Per_Deliverable_Views")
    deliverable_views = flatten_per_deliverable_views(
        data.get("per_deliverable_views_summary", {})
    )

    deliv_headers = ["deliverable", "score"]
    for col_idx, header in enumerate(deliv_headers, start=1):
        ws_deliv.cell(row=1, column=col_idx, value=header)
        ws_deliv.cell(row=1, column=col_idx).font = Font(bold=True)

    for row_idx, deliv in enumerate(deliverable_views, start=2):
        for col_idx, header in enumerate(deliv_headers, start=1):
            ws_deliv.cell(row=row_idx, column=col_idx, value=deliv.get(header))

    # ===== Sheet 4: Flags =====
    ws_flags = wb.create_sheet("Flags")
    flags = flatten_flags(data.get("all_flags_aggregated", []))

    flag_headers = ["flag_id", "flag_name", "severity", "origin_stage", "indicator_id"]
    for col_idx, header in enumerate(flag_headers, start=1):
        ws_flags.cell(row=1, column=col_idx, value=header)
        ws_flags.cell(row=1, column=col_idx).font = Font(bold=True)

    # Color code by severity
    severity_colors = {
        "CRITICAL": "FF0000",  # Red
        "HIGH": "FFA500",  # Orange
        "MEDIUM": "FFFF00",  # Yellow
        "INFORMATIONAL": "00FF00",  # Green
    }

    for row_idx, flag in enumerate(flags, start=2):
        for col_idx, header in enumerate(flag_headers, start=1):
            cell = ws_flags.cell(row=row_idx, column=col_idx, value=flag.get(header))
            # Color code severity column
            if header == "severity":
                severity = flag.get("severity", "").split()[0]  # Get first word
                color = severity_colors.get(severity, "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # ===== Sheet 5: Apex Formula & Weights =====
    ws_formula = wb.create_sheet("Apex_Formula")
    ws_formula["A1"] = "Apex Formula"
    ws_formula["A1"].font = Font(bold=True, size=12)
    ws_formula["A2"] = data.get("apex_formula_spec", "")

    ws_formula["A4"] = "Weights Used"
    ws_formula["A4"].font = Font(bold=True)
    apex_weights = data.get("apex_weights_used", {})
    for idx, (block_id, weight) in enumerate(sorted(apex_weights.items()), start=5):
        ws_formula[f"A{idx}"] = block_id
        ws_formula[f"B{idx}"] = weight
        ws_formula[f"A{idx}"].font = Font(bold=True)

    # Auto-adjust column widths
    for ws in [ws_summary, ws_contrib, ws_deliv, ws_flags, ws_formula]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)


def export_all_stages_csv(
    pipeline_outputs: dict[str, dict],
    output_path: Path,
) -> None:
    """Export all pipeline stages to CSV files.

    Creates a CSV for each stage output with filename: output_path_stage_name.csv

    Args:
        pipeline_outputs: Dict mapping stage names to envelope dicts
        output_path: Base path for output files (without extension)
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    for stage_name, envelope in pipeline_outputs.items():
        data = envelope.get("data", {})
        stage_path = output_path.parent / f"{output_path.stem}_{stage_name}.csv"

        with stage_path.open("w", newline="", encoding="utf-8") as fh:
            # Flatten the data dict to single row
            flattened = _flatten_dict(data)
            writer = csv.DictWriter(fh, fieldnames=sorted(flattened.keys()))
            writer.writeheader()
            writer.writerow(flattened)


def _flatten_dict(d: dict, parent_key: str = "", sep: str = "_") -> dict:
    """Recursively flatten nested dict."""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(_flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, (list, tuple)):
            items.append((new_key, json.dumps(v)))
        else:
            items.append((new_key, v))
    return dict(items)
